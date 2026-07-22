#!/usr/bin/env python3
"""Rebuild File Explorer manifest + in-browser file contents from public/task-files.

Usage:
  python3 scripts/build-file-explorer.py

Drop new task packs under:
  public/task-files/<task_id>/required_files/...
  public/task-files/<task_id>/distractor_files/...   (optional)

Requires: mammoth, openpyxl (pip install mammoth openpyxl)
"""

from __future__ import annotations

import json
from pathlib import Path

import mammoth
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TASK_FILES = ROOT / "public" / "task-files"
MANIFEST_OUT = ROOT / "src" / "content" / "taskFilesManifest.json"
CONTENTS_OUT = ROOT / "src" / "content" / "taskFileContents.json"

LABELS = {
    "hmda_fair_lending_review": "HMDA Fair Lending Review",
    "health_gantt_89_discharge_plan_400": "Health Gantt — Discharge Plan 400",
}


def human_label(task_id: str) -> str:
    if task_id in LABELS:
        return LABELS[task_id]
    return task_id.replace("_", " ")


def build_manifest() -> dict:
    tasks = []
    for task_dir in sorted(TASK_FILES.iterdir()):
        if not task_dir.is_dir():
            continue
        files = []
        for kind_dir_name, kind in (
            ("required_files", "required"),
            ("distractor_files", "distractor"),
        ):
            kind_dir = task_dir / kind_dir_name
            if not kind_dir.is_dir():
                continue
            for f in sorted(kind_dir.iterdir()):
                if f.name.startswith(".") or not f.is_file():
                    continue
                files.append(
                    {
                        "name": f.name,
                        "path": f.relative_to(TASK_FILES).as_posix(),
                        "kind": kind,
                        "ext": f.suffix.lstrip(".").lower(),
                        "size_bytes": f.stat().st_size,
                    }
                )
        tasks.append(
            {
                "id": task_dir.name,
                "label": human_label(task_dir.name),
                "files": files,
            }
        )
    return {"tasks": tasks}


def build_contents() -> dict:
    contents: dict = {}
    for path in sorted(TASK_FILES.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        rel = path.relative_to(TASK_FILES).as_posix()
        ext = path.suffix.lstrip(".").lower()
        if ext == "docx":
            with path.open("rb") as f:
                result = mammoth.convert_to_html(f)
            contents[rel] = {
                "kind": "html",
                "html": result.value,
                "messages": [str(m) for m in result.messages],
            }
        elif ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(path, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(["" if c is None else str(c) for c in row])
                while rows and all(c == "" for c in rows[-1]):
                    rows.pop()
                sheets.append({"name": name, "rows": rows})
            contents[rel] = {"kind": "sheets", "sheets": sheets}
        elif ext in ("csv", "gan", "txt", "xml", "json", "md"):
            contents[rel] = {
                "kind": "text",
                "text": path.read_text(encoding="utf-8", errors="replace"),
            }
        elif ext == "pdf":
            contents[rel] = {"kind": "pdf"}
        else:
            contents[rel] = {"kind": "download-only"}
    return contents


def main() -> None:
    if not TASK_FILES.is_dir():
        raise SystemExit(f"Missing {TASK_FILES}")
    manifest = build_manifest()
    contents = build_contents()
    MANIFEST_OUT.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    CONTENTS_OUT.write_text(
        json.dumps(contents, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {MANIFEST_OUT.relative_to(ROOT)} ({len(manifest['tasks'])} tasks)")
    print(f"Wrote {CONTENTS_OUT.relative_to(ROOT)} ({len(contents)} files)")


if __name__ == "__main__":
    main()
